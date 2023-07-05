from bs4 import BeautifulSoup
import requests
import sys
import openpyxl
import os


class AllStationDataGrabber:
    urlList = {
        '桃園市': 'https://e-service.cwb.gov.tw/HistoryDataQuery/DayDataController.do?command=viewMain&station=467050&stname=%25E6%2596%25B0%25E5%25B1%258B&datepicker=2022-12-01&altitude=20.6m',
        '屏東縣': 'https://e-service.cwb.gov.tw/HistoryDataQuery/DayDataController.do?command=viewMain&station=467590&stname=%25E6%2581%2586%25E6%2598%25A5&datepicker=2022-12-01&altitude=22.3m',
        '新竹縣': 'https://e-service.cwb.gov.tw/HistoryDataQuery/DayDataController.do?command=viewMain&station=467571&stname=%25E6%2596%25B0%25E7%25AB%25B9&datepicker=2022-12-13&altitude=26.9m',
        '新竹市': 'https://e-service.cwb.gov.tw/HistoryDataQuery/DayDataController.do?command=viewMain&station=C0D660&stname=%25E6%2596%25B0%25E7%25AB%25B9%25E5%25B8%2582%25E6%259D%25B1%25E5%258D%2580&datepicker=2022-12-13&altitude=65m',
        '新北市': 'https://e-service.cwb.gov.tw/HistoryDataQuery/DayDataController.do?command=viewMain&station=466880&stname=%25E6%259D%25BF%25E6%25A9%258B&datepicker=2022-12-11&altitude=9.7m',
        '台北市': 'https://e-service.cwb.gov.tw/HistoryDataQuery/DayDataController.do?command=viewMain&station=466910&stname=%25E9%259E%258D%25E9%2583%25A8&datepicker=2022-12-13&altitude=837.6m',
        '基隆市': 'https://e-service.cwb.gov.tw/HistoryDataQuery/DayDataController.do?command=viewMain&station=466940&stname=%25E5%259F%25BA%25E9%259A%2586&datepicker=2022-12-13&altitude=26.7m',
        '苗栗縣': 'https://e-service.cwb.gov.tw/HistoryDataQuery/DayDataController.do?command=viewMain&station=C0E420&stname=%25E7%25AB%25B9%25E5%258D%2597&datepicker=2022-12-13&altitude=19m',
        '台中市': 'https://e-service.cwb.gov.tw/HistoryDataQuery/DayDataController.do?command=viewMain&station=467490&stname=%25E8%2587%25BA%25E4%25B8%25AD&datepicker=2022-12-13&altitude=84.04m',
        '澎湖縣': 'https://e-service.cwb.gov.tw/HistoryDataQuery/DayDataController.do?command=viewMain&station=467300&stname=%25E6%259D%25B1%25E5%2590%2589%25E5%25B3%25B6&datepicker=2022-12-13&altitude=43.0m',
        '彰化縣': 'https://e-service.cwb.gov.tw/HistoryDataQuery/DayDataController.do?command=viewMain&station=467270&stname=%25E7%2594%25B0%25E4%25B8%25AD&datepicker=2022-12-13&altitude=48.6m',
        '南投縣': 'https://e-service.cwb.gov.tw/HistoryDataQuery/DayDataController.do?command=viewMain&station=467550&stname=%25E7%258E%2589%25E5%25B1%25B1&datepicker=2022-12-12&altitude=3844.8m',
        '雲林縣': 'https://e-service.cwb.gov.tw/HistoryDataQuery/DayDataController.do?command=viewMain&station=C0K240&stname=%25E8%258D%2589%25E5%25B6%25BA&datepicker=2022-12-13&altitude=1132m',
        '嘉義縣': 'https://e-service.cwb.gov.tw/HistoryDataQuery/DayDataController.do?command=viewMain&station=467530&stname=%25E9%2598%25BF%25E9%2587%258C%25E5%25B1%25B1&datepicker=2022-12-13&altitude=2413.4m',
        '嘉義市': 'https://e-service.cwb.gov.tw/HistoryDataQuery/DayDataController.do?command=viewMain&station=467480&stname=%25E5%2598%2589%25E7%25BE%25A9&datepicker=2022-12-13&altitude=26.9m',
        '台南市': 'https://e-service.cwb.gov.tw/HistoryDataQuery/DayDataController.do?command=viewMain&station=467410&stname=%25E8%2587%25BA%25E5%258D%2597&datepicker=2022-12-13&altitude=40.8m',
        '高雄市': 'https://e-service.cwb.gov.tw/HistoryDataQuery/DayDataController.do?command=viewMain&station=467440&stname=%25E9%25AB%2598%25E9%259B%2584&datepicker=2022-12-13&altitude=2.3m',
        '宜蘭縣': 'https://e-service.cwb.gov.tw/HistoryDataQuery/DayDataController.do?command=viewMain&station=467060&stname=%25E8%2598%2587%25E6%25BE%25B3&datepicker=2022-12-13&altitude=24.9m',
        '花蓮縣': 'https://e-service.cwb.gov.tw/HistoryDataQuery/DayDataController.do?command=viewMain&station=466990&stname=%25E8%258A%25B1%25E8%2593%25AE&datepicker=2022-12-13&altitude=16.1m',
        '台東縣': 'https://e-service.cwb.gov.tw/HistoryDataQuery/DayDataController.do?command=viewMain&station=467540&stname=%25E5%25A4%25A7%25E6%25AD%25A6&datepicker=2022-12-13&altitude=8.1m'
    }
    abandonSt = {'桃園市': ('水尾', '撤銷站'),
                 '屏東縣': ('琉球嶼', '口社', '龍泉', '西大武山', '石門山', '上德文', '力里', '九如', '瑪家', '崁頂', '里港', '竹田',  '撤銷站'),
                 '新竹縣': ('新豐', '峨眉', '新埔', '鳥嘴山', '白蘭', '太閣南', '飛鳳山', '外坪(五指山)', '撤銷站'),
                 '新竹市': ('撤銷站'),
                 '新北市': ('福隆','福山','泰平', '石碇服務區', '坪林交控', '四十份', '國一N039K', '下盆', '鼻頭角', '富貴角', '撤銷站'),
                 '台北市': ('關渡', '國三N016K', '撤銷站'),
                 '基隆市': ('彭佳嶼', '撤銷站', '基隆嶼'),
                 '苗栗縣': ('加祿堂', '大河', '獅潭', '頭份', '南庄', '竹南', '象鼻', '松安', '鳳美', '新開', '南勢', '南礦', '南勢山', '南湖', '八卦', '馬拉邦山', '泰安', '公館', '國三N149K', '國一N128K', '撤銷站'),
                 '台中市': ('雪山圈谷', '上谷關', '稍來', '新伯公', '雪嶺', '桐林', '白冷', '白毛台', '龍安', '伯公龍', '慶福山', '清水林', '德基', '撤銷站'),
                 '澎湖縣': ('撤銷站'),
                 '彰化縣': ('下水埔', '國一S218K', '撤銷站'),
                 '南投縣': ('翠峰', '國三N238K', '瑞岩', '清流', '長豐', '雙冬', '六分寮', '阿眉', '萬大', '武界', '丹大', '和社', '溪頭', '大鞍', '桶頭', '卡奈托灣', '青雲', '中心崙', '蘆竹湳', '樟湖', '九份二山', '外大坪',
                         '鯉潭', '北坑', '埔中', '豐丘', '西巒', '奧萬大', '楓樹林', '新興橋', '凌霄', '翠華', '新高口', '望鄉山', '杉林溪', '大尖山', '線浸林道', '國六W023K', '玉山風口', '奇萊稜線', '撤銷站'),
                 '雲林縣': ('口湖', '撤銷站'),
                 '嘉義縣': ('龍美', '菜瓜坪', '獨立山', '頭凍', '石磐龍', '瑞里', '十字', '國三N285K', '撤銷站'),
                 '嘉義市': ('撤銷站'),
                 '台南市': ('沙崙', '環湖', '大棟山', '關山', '楠西', '東山服務區', '東原', '撤銷站'),
                 '高雄市': ('達卡努瓦', '排雲', '南天池', '梅山', '小關山', '高中', '御油山', '大津', '尖山', '吉東', '溪南(特生中心)', '新發', '藤枝', '多納林道', '國三S383K', '林園', '大寮', '撤銷站'),
                 '宜蘭縣': ('龜山島', '牛鬥', '寒溪', '東澳嶺', '觀音海岸', '思源', '粉鳥林', '撤銷站'),
                 '花蓮縣': ('洛韶', '慈恩', '布洛灣', '中興', '大觀', '太安', '大農', '龍澗', '高寮', '太魯閣', '豐南', '紅葉', '立山', '三棧', '壽豐', '銅門', '荖溪', '中平林道', '和中', '撤銷站'),
                 '台東縣': ('蘭嶼', '蘭嶼高中', '綠島', '摩天', '華源', '金峰', '利嘉', '南美山', '壽卡', '利嘉林道', '都蘭', '蘭嶼燈塔', '撤銷站')
                 }
    urlList.pop('嘉義市')
    abandonSt.pop('嘉義市')
    urlList.pop('澎湖縣')
    abandonSt.pop('澎湖縣')
    areaStationsList = {}

    def __init__(self, app) -> None:
        self.app = app
        with open(os.path.join(self.app.datas, 'new_station.txt'), 'r') as f:
            newStationList = f.readlines()
        newStationList = [line[:-1] for line in newStationList]

        #  ---initialize---
        for city, url in self.urlList.items():
            self.areaStationsList[city] = self.__get_stations_info(city, url)
            for number in self.areaStationsList[city].copy():
                if number in newStationList:
                    self.areaStationsList[city].pop(number)
        #  ---initialize---

    def __get_stations_info(self, city, url):
        wb = openpyxl.load_workbook('datas\\station_location.xlsx')
        sheet = wb[wb.sheetnames[0]]

        request = requests.get(url)
        state = request.status_code

        if state != 200:
            print('Something Wrong')
            sys.exit()

        htmlParse = BeautifulSoup(request.text, 'html.parser')
        stations_info = {}   # {'469010': {'stName': '新屋', 'stAlti': '5m', ~~~~}}
        for td in htmlParse.find('tr').find_all('td'):
            if td.find('select'):
                for stations in td.find('select').find_all('option'):
                    for abName in AllStationDataGrabber.abandonSt[city]:
                        if abName in stations.getText():
                            break
                    else:
                        # --- name, alti ---
                        stNo, stName, stAlti = stations.get('value').split('_')
                        stations_info[str(stNo)] = {
                            'stName': stName, 'stAlti': stAlti}
                        # --- location ---
                        for value in zip(sheet['A'], sheet['D'], sheet['E']):
                            if str(value[0].value) == stNo:
                                stations_info[stNo]['stLon'] = value[1].value
                                stations_info[stNo]['stLat'] = value[2].value
                                break
        return stations_info
    '''
    def __check_new_station(self, city):
        
        for number in self.areaStationsList[city]:
            self.single_info(number, (2022, 12, 24), True, city)
    '''

    def __create_xlsx(self, stNo, temperature, rainfull):
        if not os.path.exists(os.path.join('datas', f'{self.app.select_area.get()}{self.__date}')):
            os.mkdir(os.path.join(
                'datas', f'{self.app.select_area.get()}{self.__date}'))

        wb = openpyxl.Workbook()
        sheet = wb['Sheet']
        sheet.title = self.areaStationsList[self.app.select_area.get(
        )][stNo]['stName']

        for column, value in zip(('A', 'B', 'C'), ('Hour', 'Rain', 'T')):
            sheet[f'{column}1'].value = value
        for i in range(2, 26):
            sheet[f'A{i}'].value = i-1
            sheet[f'B{i}'].value = rainfull[i-2]
            sheet[f'C{i}'].value = temperature[i-2]
        wb.save(os.path.join(
            'datas', f'{self.app.select_area.get()}{self.__date}', f'{stNo}.xlsx'))
    # -------------------------------------------------------------------------------------------

    def single_info(self, stNo, currentDate: tuple, parse=None, dir_area=None):
        '''
        passing a stNo and return temp and rainfall
        '''
        temperature, rainfall = [], []

        if parse:
            stInfo = self.areaStationsList[dir_area][stNo]
        else:
            stInfo = self.areaStationsList[self.app.select_area.get()][stNo]
        self.__date = self.app.str_tuple_date_change(currentDate)
        stName = ''.join(['%25'+(hex(hex_)[2:].upper())
                          for hex_ in stInfo['stName'].encode(encoding='utf-8')])
        stAlti = stInfo['stAlti']

        url = f'https://e-service.cwb.gov.tw/HistoryDataQuery/DayDataController.do?command=viewMain&station={stNo}&stname={stName}&datepicker={self.__date}&altitude={stAlti}'

        request = requests.get(url)
        state = request.status_code
        if state != 200:
            print('Something Wrong')
            sys.exit()

        htmlParse = BeautifulSoup(request.text, 'html.parser')

        # if parse:
        #     if htmlParse.find('label', class_='imp').getText() == '本段時間區間內無觀測資料。':
        #         with open(os.path.join(self.app.datas, 'new_station.txt'), 'a') as f:
        #             f.write(str(stNo)+'\n')

        # else:

        for tr in htmlParse.find('div', class_='CSSTableGenerator').find_all('tr'):
            if tr.find('td'):
                hour_data = tr.find_all('td')
                temperature.append(hour_data[3].getText()[:-1])
                rainfall.append(hour_data[10].getText()[:-1])

        self.__create_xlsx(stNo, temperature, rainfall)

    def check_repeat(self, date: tuple):
        for stNo in self.areaStationsList[self.app.select_area.get()]:
            if os.path.exists(os.path.join(self.app.datas, f'{self.app.select_area.get()}{self.app.str_tuple_date_change(date)}', f'{stNo}.xlsx')):
                pass
            else:
                return False
        else:
            return True


if __name__ == '__main__':
    pass
